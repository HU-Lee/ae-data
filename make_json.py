import os
import re
from openpyxl import load_workbook
import json
from config import FILE_NAME, ELEMENTS, WEAPONS

if not os.path.exists("json"):
    os.makedirs("json")

workbook = load_workbook(FILE_NAME, data_only=True)


'''
character.json 파일 만들기
'''
char_sheet = workbook["캐릭터"]

char_arr = []
codes = []


for row in char_sheet.iter_rows(min_row=2):
    dic = {
        "id": row[0].value,
        "code": str(row[3].value),
        "style": str(row[2].value).lower(),
        "category": ELEMENTS.index(row[4].value)*10 + WEAPONS.index(row[5].value),
        "free": row[6].value,
        "sky": "light" if row[7].value == "천" else "shadow",
        "first": row[3].value not in codes,
        "jonly": row[8].value,
        "gonly": row[9].value,
        "from": list(map(int, str(row[10].value).split(","))) if row[10].value else [],
        "change": list(map(int, str(row[11].value).split(","))) if row[11].value else [],
        "book": row[12].value or "없음",
        "book_get": row[13].value.split(",") if row[13].value else [],
        "manifest_jap": row[15].value or "없음",
        "manifest_glo": row[16].value or "없음"
    }
    char_arr.append(dic)
    codes = list(set(codes + [row[3].value]))

with open('json/character.json', 'w+', encoding='utf-8') as f:
    json.dump(char_arr, f, indent="\t", ensure_ascii=False)


'''
번역 json 파일 만들기
'''
kor_json = {}
eng_json = {}
jap_json = {}

# 1. 속성, 무기 번역
etc_trans_sheet = workbook["기타번역"]
for row in etc_trans_sheet.iter_rows(min_row=2):
    eng_key = re.sub(r"[^a-zA-Z0-9]","",row[1].value).lower()
    kor_json[eng_key] = row[0].value
    eng_json[eng_key] = row[1].value
    jap_json[eng_key] = row[2].value

# 2. 설명 번역
desc_trans_sheet = workbook["설명"]
for row in desc_trans_sheet.iter_rows(min_row=2):
    eng_key = row[3].value
    if not eng_key: 
        continue
    kor_json[eng_key] = row[0].value
    eng_json[eng_key] = row[1].value
    jap_json[eng_key] = row[2].value

# 3. 캐릭터 이름 번역
for row in char_sheet.iter_rows(min_row=2):
    kor_json[str(row[3].value)] = row[1].value

char_trans_sheet = workbook["캐릭번역"]
for row in char_trans_sheet.iter_rows(min_row=2):
    for key in kor_json.keys():
        if kor_json[key] == row[0].value:
            eng_json[key] = row[1].value
            jap_json[key] = row[2].value

# 4. 나머지 작업
for i in ["던전", "기타번역", "캐릭번역", "특성번역"]:
    sheet = workbook[i]
    for row in sheet.iter_rows(min_row=2):
        kor_key = row[0].value
        kor_json[kor_key] = row[0].value
        eng_json[kor_key] = row[1].value
        jap_json[kor_key] = row[2].value

with open('json/ko.json', 'w+', encoding='utf-8') as f:
    json.dump(kor_json, f, indent="\t", ensure_ascii=False)
    
with open('json/en.json', 'w+', encoding='utf-8') as f:
    json.dump(eng_json, f, indent="\t", ensure_ascii=False)
    
with open('json/jp.json', 'w+', encoding='utf-8') as f:
    json.dump(jap_json, f, indent="\t", ensure_ascii=False)



'''
dungeon.json 파일 만들기
'''
dun_arr = []

dun_sheet = workbook["던전"]
for row in dun_sheet.iter_rows(min_row=2):
    dun_arr.append({
        "name": row[0].value,
        "endpoint": row[3].value
    })

with open('json/dungeon.json', 'w+', encoding='utf-8') as f:
    json.dump(dun_arr, f, indent="\t", ensure_ascii=False)



'''
personality.json 파일 만들기
'''
per_arr = []

per_sheet = workbook["퍼스널리티"]
for row in per_sheet.iter_rows(min_row=2):
    per_arr.append({
        "name": row[0].value,
        "is_extra": row[1].value,
        "personality": row[2].value.split(","),
        "description": row[3].value,
        "code": str(row[4].value)
    })

with open('json/personality.json', 'w+', encoding='utf-8') as f:
    json.dump(per_arr, f, indent="\t", ensure_ascii=False)

'''
migration.json 파일 만들기
기존 캐시를 현재 데이터에 맞게 원복하는데 사용
'''
with open('rawdata/character_old.json', 'r', encoding='utf-8') as f:
    old_char = json.load(f)
with open('json/character.json', 'r', encoding='utf-8') as f:
    new_char = json.load(f)

# 코드를 아예 변경한 소피아와 알도는 예외처리
dic = {
    267: 254,
	268: 255,
    279: 260,
	280: 261,
}
for old in old_char:
    for new in new_char:
        if new["code"] == old["code"] and new["style"] == old["style"]:
            dic[old["id"]] = new["id"]
    if old["id"] not in dic.keys():
        dic[old["id"]] = "수정 필요"

with open('json/old_to_new.json', 'w+', encoding='utf-8') as f:
    json.dump(dic, f, indent="\t", ensure_ascii=False)